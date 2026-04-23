# ================================================
# utils/excel.py — buyurtmalarni Excel ga export
# ================================================

import logging
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db.connection import get_pool

logger = logging.getLogger(__name__)


async def export_orders_excel(limit: int = 500) -> io.BytesIO | None:
    """
    Oxirgi buyurtmalarni Excel fayliga chiqaradi.
    BytesIO qaytaradi — to'g'ridan-to'g'ri Telegram ga yuborsa bo'ladi.
    """
    try:
        pool = get_pool()
        async with pool.acquire() as conn:
            orders = await conn.fetch(
                "SELECT o.*, u.full_name, u.username, u.phone AS user_phone"
                " FROM orders o"
                " LEFT JOIN users u ON o.user_id=u.id"
                " ORDER BY o.id DESC LIMIT $1",
                limit
            )
            if not orders:
                return None

            # Har bir buyurtma uchun itemlarni olish
            order_items = {}
            for order in orders:
                items = await conn.fetch(
                    "SELECT * FROM order_items WHERE order_id=$1", order["id"]
                )
                order_items[order["id"]] = [dict(i) for i in items]

        # ── Excel yaratish ────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Buyurtmalar"

        # Ranglar
        HEADER_COLOR = "2E4057"
        ROW_ODD      = "F8F9FA"
        ROW_EVEN     = "FFFFFF"
        GREEN        = "28A745"
        RED          = "DC3545"
        YELLOW       = "FFC107"
        BLUE         = "007BFF"
        GRAY         = "6C757D"

        STATUS_COLORS = {
            "kutilmoqda":    YELLOW,
            "qabul qilindi": BLUE,
            "yo'lda":        "17A2B8",
            "yetkazildi":    GREEN,
            "bekor qilindi": RED,
        }

        def header_style(cell, text):
            cell.value = text
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = Border(
                bottom=Side(style="medium", color="FFFFFF")
            )

        def cell_style(cell, value, color=None, bold=False, center=False):
            cell.value     = value
            cell.font      = Font(bold=bold, size=10)
            cell.alignment = Alignment(
                horizontal="center" if center else "left",
                vertical="center", wrap_text=True
            )
            if color:
                cell.fill = PatternFill("solid", fgColor=color)

        # ── Sarlavha ──────────────────────────────────
        ws.merge_cells("A1:K1")
        title_cell = ws["A1"]
        title_cell.value     = f"🌸 Sifat Parfimer Shop — Buyurtmalar ({datetime.now().strftime('%d.%m.%Y')})"
        title_cell.font      = Font(bold=True, size=14, color=HEADER_COLOR)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # ── Ustun nomlari ─────────────────────────────
        headers = [
            "№", "Sana", "Mijoz", "Telefon",
            "Manzil", "Mahsulotlar", "Izoh",
            "Summa", "Yetkazish", "Jami", "Holat"
        ]
        for col, h in enumerate(headers, 1):
            header_style(ws.cell(row=2, column=col), h)
        ws.row_dimensions[2].height = 25

        # ── Ma'lumotlar ───────────────────────────────
        for row_idx, order in enumerate(orders, 3):
            o       = dict(order)
            items   = order_items.get(o["id"], [])
            bg      = ROW_ODD if row_idx % 2 == 0 else ROW_EVEN
            status  = o.get("status", "")
            st_color = STATUS_COLORS.get(status, GRAY)

            # Mahsulotlar matni
            prod_text = "\n".join(
                f"• {i['name']}"
                + (f" ({i['variant_name']})" if i.get("variant_name") else "")
                + f" × {i['qty']} = {i['price'] * i['qty']:,} so'm"
                for i in items
            )

            # Sana
            created = str(o.get("created_at", ""))[:16]

            # Jami
            total    = o.get("total", 0)
            delivery = o.get("delivery_price", 0)
            grand    = total + delivery

            values = [
                o["id"],
                created,
                o.get("full_name") or "—",
                o.get("phone", "—"),
                o.get("address") or "—",
                prod_text,
                o.get("comment") or "—",
                f"{total:,} so'm",
                f"{delivery:,} so'm" if delivery else "Bepul",
                f"{grand:,} so'm",
                status,
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col)
                is_status = col == 11
                cell_style(
                    cell, val,
                    color=st_color if is_status else bg,
                    bold=is_status,
                    center=(col in (1, 2, 8, 9, 10, 11))
                )
                if is_status:
                    cell.font = Font(bold=True, color="FFFFFF", size=10)

            ws.row_dimensions[row_idx].height = max(30, len(items) * 18)

        # ── Ustun kengliklari ─────────────────────────
        col_widths = [6, 16, 20, 16, 22, 45, 20, 16, 14, 16, 16]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # ── Statistika varag'i ────────────────────────
        ws2 = wb.create_sheet("Statistika")

        total_orders   = len(orders)
        total_revenue  = sum(dict(o)["total"] for o in orders
                             if dict(o)["status"] != "bekor qilindi")
        total_delivery = sum(dict(o).get("delivery_price", 0) for o in orders
                             if dict(o)["status"] != "bekor qilindi")
        cancelled      = sum(1 for o in orders
                             if dict(o)["status"] == "bekor qilindi")
        delivered      = sum(1 for o in orders
                             if dict(o)["status"] == "yetkazildi")

        stats = [
            ("📦 Jami buyurtmalar",        total_orders),
            ("✅ Yetkazilgan",              delivered),
            ("❌ Bekor qilingan",           cancelled),
            ("💰 Jami daromad",            f"{total_revenue:,} so'm"),
            ("🚚 Jami yetkazib berish",    f"{total_delivery:,} so'm"),
            ("💳 Umumiy jami",             f"{total_revenue + total_delivery:,} so'm"),
        ]

        ws2["A1"].value     = "📊 Statistika"
        ws2["A1"].font      = Font(bold=True, size=14, color=HEADER_COLOR)
        ws2["A1"].alignment = Alignment(horizontal="center")
        ws2.merge_cells("A1:B1")
        ws2.row_dimensions[1].height = 28

        for i, (label, val) in enumerate(stats, 2):
            ws2.cell(row=i, column=1).value = label
            ws2.cell(row=i, column=1).font  = Font(bold=True, size=11)
            ws2.cell(row=i, column=2).value = val
            ws2.cell(row=i, column=2).font  = Font(size=11)
            ws2.cell(row=i, column=2).alignment = Alignment(horizontal="right")
            ws2.row_dimensions[i].height = 22

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 20

        # ── BytesIO ga saqlash ────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    except Exception as e:
        logger.error(f"export_orders_excel: {e}"); return None
