# ================================================
# utils/backup.py — butun bazaning to'liq bekap nusxasi
# ================================================
# Har bir jadval (categories, products, users, orders va h.k.)
# alohida Excel varag'iga yoziladi. Agar baza qandaydir sababdan
# yo'qolib qolsa, shu fayldagi ma'lumotlardan qo'lda qayta tiklash
# mumkin bo'ladi.

import logging
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from db.connection import get_pool

logger = logging.getLogger(__name__)

# Bekapga kiritiladigan jadvallar — db/init_db.py dagi CREATE TABLE
# ro'yxati bilan bir xil bo'lishi kerak
BACKUP_TABLES = [
    "categories", "subcategories", "products", "product_photos",
    "product_variants", "users", "orders", "order_items",
    "carts", "favorites", "last_seen",
]

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(color="FFFFFF", bold=True)


async def export_full_backup() -> io.BytesIO | None:
    """
    Barcha jadvallarni bitta Excel faylga, har birini alohida
    varaqqa yozib chiqaradi. Xato chiqsa ham, qaysi jadvallar
    muvaffaqiyatli o'qilgan bo'lsa, ular saqlanadi — bitta jadvalda
    muammo bo'lsa ham, butun bekap ishi to'xtab qolmaydi.
    """
    try:
        pool = get_pool()
        wb = Workbook()
        wb.remove(wb.active)  # standart bo'sh varaqni o'chirish

        any_data = False
        async with pool.acquire() as conn:
            for table in BACKUP_TABLES:
                try:
                    rows = await conn.fetch(f"SELECT * FROM {table} ORDER BY 1")
                except Exception as e:
                    logger.error(f"export_full_backup: '{table}' jadvalini o'qishda xato: {e}")
                    continue

                ws = wb.create_sheet(title=table[:31])  # Excel varaq nomi <=31 belgi

                if not rows:
                    ws.append(["(bo'sh)"])
                    continue

                any_data = True
                columns = list(rows[0].keys())
                ws.append(columns)
                for col_idx in range(1, len(columns) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center")

                for row in rows:
                    values = []
                    for v in row.values():
                        # datetime va boshqa maxsus turlarni Excel
                        # tushunadigan oddiy matn/raqamga aylantirish
                        if isinstance(v, datetime):
                            values.append(v.strftime("%Y-%m-%d %H:%M:%S"))
                        elif v is None:
                            values.append("")
                        else:
                            values.append(str(v) if not isinstance(v, (int, float, bool)) else v)
                    ws.append(values)

                for col_idx, col_name in enumerate(columns, 1):
                    max_len = max(
                        [len(str(col_name))] +
                        [len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)]
                    )
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        if not any_data:
            return None

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    except Exception as e:
        logger.error(f"export_full_backup: umumiy xato: {e}")
        return None
